import os
import json
import openpyxl
import pytz
from datetime import datetime, timezone
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
from apscheduler.schedulers.background import BackgroundScheduler

from broadcast_handler import init_broadcast, do_scheduled_broadcast, restore_scheduled_jobs
from scenario_handler import init_scenarios

BOT_TOKEN = os.environ.get("BOT_TOKEN")
# Читаем список администраторов из переменной окружения ADMIN_IDS
ADMIN_IDS = [int(x) for x in os.environ.get("ADMIN_IDS", "").split(",") if x]

bot = __import__("telebot").TeleBot(BOT_TOKEN)
scheduler = BackgroundScheduler()
scheduler.start()

# Создаем необходимые файлы, если их нет
FILE_LIST = {
    "scenario_store.json": {},
    "user_db.json": [],
    "broadcasts.json": {},
    "scheduled_broadcasts.json": [],
    "temp_broadcasts.json": {},
    "temp_scenarios.json": {}
}
for filename, default_data in FILE_LIST.items():
    if not os.path.exists(filename):
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(default_data, f, ensure_ascii=False)

def update_user_activity(user):
    try:
        with open("user_db.json", "r", encoding="utf-8") as f:
            users = json.load(f)
    except Exception:
        users = []
    found = False
    for u in users:
        if u["id"] == user.id:
            u["last_active"] = datetime.now(timezone.utc).isoformat()
            found = True
            break
    if not found:
        new_user = {
            "id": user.id,
            "first_name": getattr(user, "first_name", ""),
            "username": getattr(user, "username", ""),
            "phone": "",
            "last_active": datetime.now(timezone.utc).isoformat()
        }
        users.append(new_user)
    with open("user_db.json", "w", encoding="utf-8") as f:
        json.dump(users, f, ensure_ascii=False)

def send_content(chat_id, text, file_id=None, link=None, file_type=None):
    final_text = text
    if link:
        final_text += f"\n\n🔗 {link}"
    try:
        if file_id:
            if file_type == "photo":
                bot.send_photo(chat_id, file_id, caption=final_text)
            elif file_type == "video":
                bot.send_video(chat_id, file_id, caption=final_text)
            elif file_type == "audio":
                bot.send_audio(chat_id, file_id, caption=final_text)
            elif file_type == "animation":
                bot.send_animation(chat_id, file_id, caption=final_text)
            else:
                bot.send_document(chat_id, file_id, caption=final_text)
        else:
            bot.send_message(chat_id, final_text)
    except Exception as e:
        bot.send_message(chat_id, f"❌ Ошибка при отправке: {e}")

@bot.message_handler(commands=["start"])
def handle_start(message):
    update_user_activity(message.from_user)
    args = message.text.split()
    if len(args) > 1:
        scenario_code = args[1]
        try:
            with open("scenario_store.json", "r", encoding="utf-8") as f:
                scenarios = json.load(f)
        except:
            scenarios = {}
        scenario = scenarios.get(scenario_code)
        if scenario:
            send_content(
                chat_id=message.chat.id,
                text=scenario["text"],
                file_id=scenario.get("file_id"),
                link=scenario.get("link"),
                file_type=scenario.get("file_type")
            )
            return
        else:
            bot.send_message(message.chat.id, "❌ Такой сценарий не найден.")
    bot.send_message(message.chat.id, (
        "Привет! Добро пожаловать в бот, это тест!.\n\n"
        "/контакты – выгрузка контактов (Excel) с ссылками на диалог.\n"
        "/пользователи – статистика пользователей (общее и активные за 7 дней).\n"
        "/рассылка – создание и отправка рассылки.\n"
        "/сценарий – создание сценария (для интерактивного запуска через /start <код>).\n"
        "/скачать_сценарии_excel – экспорт сценариев в Excel.\n"
        "/скачать_рассылки_excel – экспорт рассылок в Excel.\n"
    ))

@bot.message_handler(commands=["контакты"])
def handle_contacts(message):
    if message.from_user.id not in ADMIN_IDS:
        return
    try:
        with open("user_db.json", "r", encoding="utf-8") as f:
            users = json.load(f)
    except:
        users = []
    if not users:
        bot.send_message(message.chat.id, "Пользователей пока нет.")
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append(["ID", "Имя", "Username", "Последняя активность", "Ссылка на диалог"])
    for u in users:
        user_id = u["id"]
        username = u.get("username", "")
        if username:
            link = f"https://t.me/{username}"
        else:
            link = f"tg://user?id={user_id}"
        ws.append([user_id, u.get("first_name", ""), username, u.get("last_active", ""), link])
    wb.save("contacts.xlsx")
    with open("contacts.xlsx", "rb") as doc:
        bot.send_document(message.chat.id, doc, caption="📋 Контакты пользователей (Excel)")

@bot.message_handler(commands=["пользователи"])
def handle_users(message):
    if message.from_user.id not in ADMIN_IDS:
        return
    try:
        with open("user_db.json", "r", encoding="utf-8") as f:
            users = json.load(f)
    except:
        users = []
    total = len(users)
    active = 0
    now = datetime.now(timezone.utc)
    for u in users:
        if "last_active" in u:
            try:
                last = datetime.fromisoformat(u["last_active"])
                if (now - last).days < 7:
                    active += 1
            except:
                pass
    bot.send_message(message.chat.id, f"Всего пользователей: {total}\nАктивных (за 7 дней): {active}")

@bot.message_handler(commands=["скачать_сценарии_excel"])
def download_scenarios_excel(message):
    if message.from_user.id not in ADMIN_IDS:
        return
    try:
        with open("scenario_store.json", "r", encoding="utf-8") as f:
            scenarios = json.load(f)
    except Exception as e:
        bot.send_message(message.chat.id, f"Ошибка при чтении сценариев: {e}")
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сценарии"
    ws.append(["Код", "Текст", "Файл ID", "Тип файла", "Ссылка"])
    for code, data in scenarios.items():
        ws.append([code, data.get("text", ""), data.get("file_id", ""), data.get("file_type", ""), data.get("link", "")])
    excel_file = "scenarios.xlsx"
    wb.save(excel_file)
    with open(excel_file, "rb") as doc:
        bot.send_document(message.chat.id, doc, caption="Сценарии (Excel)")

@bot.message_handler(commands=["скачать_рассылки_excel"])
def download_broadcasts_excel(message):
    if message.from_user.id not in ADMIN_IDS:
        return
    try:
        with open("broadcasts.json", "r", encoding="utf-8") as f:
            broadcasts = json.load(f)
    except Exception as e:
        bot.send_message(message.chat.id, f"Ошибка при чтении рассылок: {e}")
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Рассылки"
    ws.append(["ID", "Текст", "Файл ID", "Тип файла", "Ссылка", "Доставлено"])
    for broadcast_id, data in broadcasts.items():
        ws.append([broadcast_id, data.get("text", ""), data.get("file_id", ""), data.get("media_type", ""), data.get("link", ""), data.get("delivered", 0)])
    excel_file = "broadcasts.xlsx"
    wb.save(excel_file)
    with open(excel_file, "rb") as doc:
        bot.send_document(message.chat.id, doc, caption="Рассылки (Excel)")

@bot.message_handler(commands=["команды"])
def admin_commands(message):
    if message.from_user.id not in ADMIN_IDS:
        return
    commands_text = (
        "/контакты – выгрузка контактов (Excel) с ссылками на диалог.\n"
        "/пользователи – статистика пользователей (общее и активные за 7 дней).\n"
        "/рассылка – создание и отправка рассылки.\n"
        "/сценарий – создание сценария (для интерактивного запуска через /start <код>).\n"
        "/скачать_сценарии_excel – экспорт сценариев в Excel.\n"
        "/скачать_рассылки_excel – экспорт рассылок в Excel.\n"
    )
    bot.send_message(message.chat.id, commands_text)

def send_weekly_statistics():
    try:
        with open("user_db.json", "r", encoding="utf-8") as f:
            users = json.load(f)
    except:
        users = []
    total = len(users)
    now = datetime.now(timezone.utc)
    new_users = sum(1 for u in users if "last_active" in u and (now - datetime.fromisoformat(u["last_active"])).days < 7)
    stats_text = (
        f"📊 Статистика за неделю:\n"
        f"Новых пользователей: {new_users}\n"
        f"Общее количество пользователей: {total}"
    )
    bot.send_message(ADMIN_IDS[0], stats_text)  # уведомляем первого админа

# Планируем еженедельную статистику (каждый понедельник 09:00 UTC)
scheduler.add_job(send_weekly_statistics, 'cron', day_of_week='mon', hour=9, minute=0)

restore_scheduled_jobs(scheduler, bot)
init_broadcast(bot, ADMIN_IDS, scheduler)
init_scenarios(bot, ADMIN_IDS)

bot.polling(none_stop=True)

